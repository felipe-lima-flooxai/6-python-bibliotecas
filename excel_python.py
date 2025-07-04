from openpyxl import Workbook, load_workbook

wb = Workbook()
ws = wb.active

ws['A1'] = 'Nome'
ws['B1'] = 'Idade'

dados = [
    ['Alice', 30],
    ['Bruno', 25],
    ['Carlos', 35]
]

for i, (nome, idade) in enumerate(dados, start=2):
    ws[f'A{i}'] = nome
    ws[f'B{i}'] = idade

wb.save('exemplo.xlsx')

wb_lido = load_workbook('exemplo.xlsx')
ws_lido = wb_lido.active

for row in ws_lido.iter_rows(min_row=1, max_col=2, values_only=True):
    print(row)