from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
#achei meio chato. Bem mais legal o do PDF.

wb = Workbook()

# Planilha de funcionários
ws = wb.active
ws.title = "Funcionários"
ws.append(["ID", "Nome", "Cargo", "Salário"])

funcionarios = [
    [1, "Ana", "Analista", 5000],
    [2, "Bruno", "Desenvolvedor", 6500],
    [3, "Carla", "Gerente", 8000]
]

for f in funcionarios:
    ws.append(f)

# Aplicar negrito no cabeçalho
for cell in ws[1]:
    cell.font = Font(bold=True)

# Criar outra planilha com fórmula
ws2 = wb.create_sheet("Resumo")
ws2["A1"] = "Total de Funcionários"
ws2["B1"] = f"=COUNTA(Funcionários!A2:A100)"
ws2["A2"] = "Soma dos Salários"
ws2["B2"] = f"=SUM(Funcionários!D2:D100)"


wb.save("empresa.xlsx")


wb_lido = load_workbook("empresa.xlsx", data_only=True)
ws_lido = wb_lido["Funcionários"]

for row in ws_lido.iter_rows(min_row=2, values_only=True):
    print(row)